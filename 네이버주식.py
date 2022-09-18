import requests
from bs4 import BeautifulSoup
import openpyxl
wb = openpyxl.Workbook()
ws = wb.create_sheet('stock')
company = ['005930', '035720', '207940']
ws['A1']='회사코드'
ws['B1']='주가'
row=2
for c in company:
    
    url = 'https://finance.naver.com/item/sise.naver?code='+c
    res = requests.get(url)
    soup = BeautifulSoup(res.text, 'html.parser')
    data = soup.select('#_nowVal') #id로 검색시 #을 붙여준다 ex)#(id태크내용)

    for i in data:
        price = i.text
        price = price.replace(',','')
        print(int(price))

    ws['A'+str(row)]=c
    ws['B'+str(row)]=price
    row=row+1

wb.save('주가저장.xlsx')